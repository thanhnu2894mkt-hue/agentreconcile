"""Buoc 3 & 4: dung workbook ket qua (5 sheet) tu ket qua doi soat.

build_report(tc, rf, summ, recon, output_path)
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
base_font = Font(name=FONT, size=10)
bold_font = Font(name=FONT, size=10, bold=True)
blue_fill = PatternFill("solid", fgColor="BDD7EE")
red_fill = PatternFill("solid", fgColor="F4B7B7")
green_fill = PatternFill("solid", fgColor="C6E0B4")
title_font = Font(name=FONT, bold=True, size=14, color="1F4E78")
sub_font = Font(name=FONT, italic=True, size=9, color="808080")
new_hdr_fill = PatternFill("solid", fgColor="C55A11")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
INT = "#,##0;(#,##0);-"
MON = "#,##0.00;(#,##0.00);-"
PCT = "0.000%"
TEXTCOLS = {"Bank Trans ID", "Refund Bank Trans ID", "Trace No", "traceNo", "Bin No", "L4 Card No", "F6 Card No"}
NEWCOLS_DISPLAY = {"ISONUS", "CHECKISONUS", "BinCountry", "BinIssuer", "BINCOUNTRY", "BINISSER",
                   "Ten ngan hang", "Feerate", "Feerate dieu chinh", "Ky chot (BookDate)", "BOOK_DATE",
                   "Ky doi soat", "Discount", "Discount_ABS", "DISCOUNT", "CL", "Ty le", "Ghi chu"}


def _S(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    s = str(x)
    return s[:-2] if s.endswith(".0") else s


def _style_header(ws, row, cols):
    names = [None] * cols if isinstance(cols, int) else list(cols)
    for c, name in enumerate(names, 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = new_hdr_fill if (name in NEWCOLS_DISPLAY) else hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _write_df(ws, df, start, fmt_map):
    cols = list(df.columns)
    for j, c in enumerate(cols, 1):
        ws.cell(row=start, column=j, value=c)
    _style_header(ws, start, cols)
    for i, (_, r) in enumerate(df.iterrows(), start + 1):
        for j, c in enumerate(cols, 1):
            v = r[c]
            cell = ws.cell(row=i, column=j)
            if c in TEXTCOLS:
                cell.value = _S(v); cell.number_format = "@"
            else:
                if isinstance(v, (np.integer,)):
                    v = int(v)
                elif isinstance(v, (np.floating,)):
                    v = None if pd.isna(v) else float(v)
                elif pd.isna(v):
                    v = None
                cell.value = v
                if c in fmt_map:
                    cell.number_format = fmt_map[c]
            cell.font = base_font; cell.border = border
    return start + len(df)


def build_report(tc, rf, summ, recon, output_path):
    wb = Workbook()

    # ---- Sheet 1: ZION_THANH_CONG ----
    ws = wb.active; ws.title = "ZION_THANH_CONG"
    cols1 = ["Bank Trans ID", "Trace No", "Sub Trans Type", "Amount", "Trans Time", "Card Type",
             "Is On Us", "ISONUS", "CHECKISONUS", "BinCountry", "BinIssuer", "Ten ngan hang",
             "Bin No", "FEERATE", "FEERATE_DC", "KY_CHOT", "MM", "DISCOUNT", "DISCOUNT_ABS", "CL", "TYLE", "Ghi chu"]
    d1 = tc.copy(); d1["Is On Us"] = tc["Is On Us"]
    d1 = d1[cols1].rename(columns={"FEERATE": "Feerate", "FEERATE_DC": "Feerate dieu chinh",
                                   "KY_CHOT": "Ky chot (BookDate)", "MM": "Ky doi soat",
                                   "DISCOUNT_ABS": "Discount_ABS", "TYLE": "Ty le"})
    ws["A1"] = "FILE ZION THANH CONG — Da dieu chinh ISONUS & FEERATE"; ws["A1"].font = title_font
    ws["A2"] = "Tieu de CAM = cot tao moi | Dong xanh = ApplePay | Dong do = JCB / Miss Feerate"; ws["A2"].font = sub_font
    fmt1 = {"Amount": INT, "Feerate": MON, "Feerate dieu chinh": MON, "Discount": INT,
            "Discount_ABS": INT, "CL": MON, "Ty le": PCT, "Ky chot (BookDate)": "0"}
    _write_df(ws, d1, 4, fmt1)
    for idx, (_, r) in enumerate(tc.iterrows(), 5):
        fill = red_fill if r["Ghi chu"] == "Miss Feerate" else (blue_fill if r["MARK"] == "AP" else (red_fill if r["MARK"] == "JCB" else None))
        if fill:
            for c in range(1, len(cols1) + 1):
                ws.cell(row=idx, column=c).fill = fill
    ws.freeze_panes = "A5"

    # ---- Sheet 2: ZION_HOAN_TIEN ----
    ws2 = wb.create_sheet("ZION_HOAN_TIEN")
    cols2 = ["Refund Bank Trans ID", "traceNo", "Sub Trans Type", "Amount", "Refund Amount", "Refund Trans Time",
             "Card Scheme", "Is On Us", "ISONUS", "CHECKISONUS", "BinCountry", "BinIssuer", "Ten ngan hang",
             "FEERATE", "FEERATE_DC", "KY_CHOT", "MM", "DISCOUNT", "DISCOUNT_ABS", "CL", "TYLE"]
    d2 = rf.copy()[cols2].rename(columns={"FEERATE": "Feerate goc", "FEERATE_DC": "Feerate dieu chinh",
                                          "KY_CHOT": "Ky chot (BookDate)", "MM": "Ky doi soat",
                                          "DISCOUNT_ABS": "Discount_ABS", "TYLE": "Ty le"})
    ws2["A1"] = "FILE ZION HOAN TIEN — DPAN: Feerate dieu chinh = AmountRefund x rate%"; ws2["A1"].font = title_font
    ws2["A2"] = "Tieu de CAM = cot tao moi | Dong xanh = ApplePay | Dong do = JCB | Dong xanh la = ABNORMAL (PHI ZLP tu HOANCV)"; ws2["A2"].font = sub_font
    fmt2 = {"Amount": INT, "Refund Amount": INT, "Feerate goc": MON, "Feerate dieu chinh": MON,
            "Discount": INT, "Discount_ABS": INT, "CL": MON, "Ty le": PCT, "Ky chot (BookDate)": "0"}
    _write_df(ws2, d2, 4, fmt2)
    for idx, (_, r) in enumerate(rf.iterrows(), 5):
        fill = blue_fill if r["MARK"] == "AP" else (red_fill if r["MARK"] == "JCB" else (green_fill if r["MARK"] == "GREEN" else None))
        if fill:
            for c in range(1, len(cols2) + 1):
                ws2.cell(row=idx, column=c).fill = fill
    ws2.freeze_panes = "A5"

    # ---- Sheet 3: BANG_TONG_HOP ----
    ws3 = wb.create_sheet("BANG_TONG_HOP")
    ws3["A1"] = f"BANG TONG HOP DOI SOAT SACOMBANK vs ZION — Ky doi soat: Thang {recon}"; ws3["A1"].font = title_font
    ws3["A2"] = "Cac GD sau ky doi soat se chot ky sau. Phi so sanh = Feerate dieu chinh (chua gom phi xu ly)."; ws3["A2"].font = sub_font
    rows = [
        ("", "So luong GD", "So tien", "So phi"),
        ("ZION — Thanh cong", summ["tc_cnt"], summ["tc_amt"], summ["tc_fee"]),
        ("ZION — Hoan tien", summ["rf_cnt"], summ["rf_amt"], summ["rf_fee"]),
        ("NET ZION (TC - RF)", None, None, None),
        ("", None, None, None),
        ("SACOMBANK — Thanh cong", summ["tc_cnt"], summ["tc_amt"], summ["sb_tc_fee"]),
        ("SACOMBANK — Hoan tien", summ["rf_cnt"], summ["rf_amt"], summ["sb_rf_fee"]),
        ("NET SACOMBANK (TC - RF)", None, None, None),
        ("", None, None, None),
        ("CHENH LECH (NET ZION - NET SACOMBANK)", None, None, None),
        ("", None, None, None),
        ("Phi xu ly GD = (TC + RF) x 440", None, None, None),
    ]
    r0 = 4
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            c = ws3.cell(row=r0 + i, column=j + 1, value=val)
            c.font = base_font; c.border = border
            if j >= 1 and isinstance(val, (int, float)) and not isinstance(val, bool):
                c.number_format = INT if float(val).is_integer() else MON
    _style_header(ws3, r0, 4)
    ws3[f"D{r0+3}"] = f"=D{r0+1}-D{r0+2}"; ws3[f"C{r0+3}"] = f"=C{r0+1}-C{r0+2}"; ws3[f"B{r0+3}"] = f"=B{r0+1}-B{r0+2}"
    ws3[f"D{r0+7}"] = f"=D{r0+5}-D{r0+6}"; ws3[f"C{r0+7}"] = f"=C{r0+5}-C{r0+6}"; ws3[f"B{r0+7}"] = f"=B{r0+5}-B{r0+6}"
    ws3[f"D{r0+9}"] = f"=D{r0+3}-D{r0+7}"; ws3[f"C{r0+9}"] = f"=C{r0+3}-C{r0+7}"; ws3[f"B{r0+9}"] = f"=B{r0+3}-B{r0+7}"
    ws3[f"D{r0+11}"] = f"=(B{r0+1}+B{r0+2})*440"
    for rr in (r0 + 3, r0 + 7, r0 + 9, r0 + 11):
        for cc in range(1, 5):
            ws3.cell(row=rr, column=cc).font = bold_font
            ws3.cell(row=rr, column=cc).number_format = INT
            if rr == r0 + 9:
                ws3.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor="FFF2CC")

    ir = r0 + 14
    ws3.cell(row=ir, column=1, value="XU LY ISONUS").font = title_font
    for j, h in enumerate(["Phan loai", "TC (so luong)", "RF (so luong)"], 1):
        ws3.cell(row=ir + 1, column=j, value=h)
    _style_header(ws3, ir + 1, 3)
    tc_cnt = tc["CHECKISONUS"].value_counts().to_dict()
    rf_cnt = rf["CHECKISONUS"].value_counts().to_dict()
    cats = sorted(set(tc_cnt) | set(rf_cnt))
    for k, cat in enumerate(cats, 1):
        ws3.cell(row=ir + 1 + k, column=1, value=cat).font = base_font
        ws3.cell(row=ir + 1 + k, column=2, value=int(tc_cnt.get(cat, 0))).font = base_font
        ws3.cell(row=ir + 1 + k, column=3, value=int(rf_cnt.get(cat, 0))).font = base_font
        for cc in range(1, 4):
            ws3.cell(row=ir + 1 + k, column=cc).border = border
    tot = ir + 1 + len(cats) + 1
    first, last = ir + 2, ir + 1 + len(cats)
    ws3.cell(row=tot, column=1, value="Tổng cộng")
    ws3.cell(row=tot, column=2, value=f"=SUM(B{first}:B{last})")
    ws3.cell(row=tot, column=3, value=f"=SUM(C{first}:C{last})")
    for cc in range(1, 4):
        cell = ws3.cell(row=tot, column=cc)
        cell.font = bold_font; cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        if cc >= 2:
            cell.number_format = INT
    ws3.column_dimensions["A"].width = 42
    for col in "BCD":
        ws3.column_dimensions[col].width = 18

    # ---- Sheet 4: TC detail ----
    tcd = tc[(tc["MM"] == recon) & (tc["CL"].round(0) != 0)].copy()
    tcd_out = pd.DataFrame({
        "Bank Trans ID": tcd["Bank Trans ID"], "Sub Trans Type": tcd["Sub Trans Type"],
        "Amount": tcd["Amount"], "Trans Time": tcd["Trans Time"], "Card Type": tcd["Card Type"],
        "Is On Us": tcd["ISONUS"], "BINCOUNTRY": tcd["BinCountry"], "BINISSER": tcd["BinIssuer"],
        "DISCOUNT": tcd["DISCOUNT_ABS"], "Feerate dieu chinh": tcd["FEERATE_DC"]})
    ws4 = wb.create_sheet("TC")
    ws4["A1"] = f"Sheet TC — GD Thanh cong CHENH LECH phi (ky {recon}, CL != 0)"; ws4["A1"].font = title_font
    ws4["A2"] = f"So GD chenh lech: {len(tcd_out)}  |  Ty le = Feerate dieu chinh / Amount"; ws4["A2"].font = sub_font
    cols4 = list(tcd_out.columns) + ["Ty le", "CL"]
    for j, c in enumerate(cols4, 1):
        ws4.cell(row=4, column=j, value=c)
    _style_header(ws4, 4, cols4)
    for i, (_, r) in enumerate(tcd_out.iterrows(), 5):
        for j, c in enumerate(tcd_out.columns, 1):
            cell = ws4.cell(row=i, column=j)
            if c in TEXTCOLS:
                cell.value = _S(r[c]); cell.number_format = "@"
            else:
                v = r[c]; v = None if (isinstance(v, float) and pd.isna(v)) else (float(v) if isinstance(v, np.floating) else (int(v) if isinstance(v, np.integer) else v))
                cell.value = v
            cell.font = base_font; cell.border = border
        ws4.cell(row=i, column=3).number_format = INT
        ws4.cell(row=i, column=9).number_format = INT
        ws4.cell(row=i, column=10).number_format = MON
        ws4.cell(row=i, column=11, value=f"=IF(C{i}=0,0,J{i}/C{i})").number_format = PCT
        ws4.cell(row=i, column=12, value=f"=J{i}-I{i}").number_format = MON
        ws4.cell(row=i, column=11).font = base_font; ws4.cell(row=i, column=12).font = base_font
        ws4.cell(row=i, column=11).border = border; ws4.cell(row=i, column=12).border = border
    ws4.freeze_panes = "A5"

    # ---- Sheet 5: RF detail ----
    rfd = rf[(rf["MM"] == recon) & (rf["CL"].round(0) != 0)].copy()
    rfd_out = pd.DataFrame({
        "Refund Bank Trans ID": rfd["Refund Bank Trans ID"], "Refund Amount": rfd["Refund Amount"],
        "Is On Us": rfd["ISONUS"], "Card Scheme": rfd["Card Scheme"], "BINCOUNTRY": rfd["BinCountry"],
        "BINISSER": rfd["BinIssuer"], "BOOK_DATE": rfd["KY_CHOT"], "DISCOUNT": rfd["DISCOUNT_ABS"],
        "Feerate": rfd["FEERATE_DC"]})
    ws5 = wb.create_sheet("RF")
    ws5["A1"] = f"Sheet RF — GD Hoan tien CHENH LECH phi (ky {recon}, CL != 0)"; ws5["A1"].font = title_font
    ws5["A2"] = f"So GD chenh lech: {len(rfd_out)}  |  Ty le = Feerate / Refund Amount"; ws5["A2"].font = sub_font
    cols5 = list(rfd_out.columns) + ["Ty le", "CL"]
    for j, c in enumerate(cols5, 1):
        ws5.cell(row=4, column=j, value=c)
    _style_header(ws5, 4, cols5)
    for i, (_, r) in enumerate(rfd_out.iterrows(), 5):
        for j, c in enumerate(rfd_out.columns, 1):
            cell = ws5.cell(row=i, column=j)
            if c in TEXTCOLS:
                cell.value = _S(r[c]); cell.number_format = "@"
            else:
                v = r[c]; v = None if (isinstance(v, float) and pd.isna(v)) else (float(v) if isinstance(v, np.floating) else (int(v) if isinstance(v, np.integer) else v))
                cell.value = v
            cell.font = base_font; cell.border = border
        ws5.cell(row=i, column=2).number_format = INT
        ws5.cell(row=i, column=7).number_format = "0"
        ws5.cell(row=i, column=8).number_format = INT
        ws5.cell(row=i, column=9).number_format = MON
        ws5.cell(row=i, column=10, value=f"=IF(B{i}=0,0,I{i}/B{i})").number_format = PCT
        ws5.cell(row=i, column=11, value=f"=I{i}-H{i}").number_format = MON
        ws5.cell(row=i, column=10).font = base_font; ws5.cell(row=i, column=11).font = base_font
        ws5.cell(row=i, column=10).border = border; ws5.cell(row=i, column=11).border = border
    ws5.freeze_panes = "A5"

    for ws_ in (ws, ws2, ws4, ws5):
        for col_cells in ws_.iter_cols(min_row=4, max_row=4):
            for cell in col_cells:
                ws_.column_dimensions[cell.column_letter].width = 15
        ws_.column_dimensions["A"].width = 22

    # Buoc cong thuc tu tinh lai khi mo file (khong can LibreOffice)
    wb.calculation.fullCalcOnLoad = True
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
